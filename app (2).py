from flask import Flask, request, send_file
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import pandas as pd
import tempfile
import os

app = Flask(__name__)

@app.route('/', methods=['POST'])
def upload_files():
    file1 = request.files['file1']
    file2 = request.files['file2']
    
    # Create temporary files
    with tempfile.NamedTemporaryFile(delete=False) as temp_file1, tempfile.NamedTemporaryFile(delete=False) as temp_file2:
        temp_file1.write(file1.read())
        temp_file2.write(file2.read())
        temp_file1_path = temp_file1.name
        temp_file2_path = temp_file2.name

    # Process the files
    result_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    result_file_path = result_file.name
    process_files(temp_file1_path, temp_file2_path, result_file_path)

    # Clean up temporary files
    os.remove(temp_file1_path)
    os.remove(temp_file2_path)

    return send_file(result_file_path, as_attachment=True)


if __name__ == '__main__':
    app.run(debug=True)


def process_files(file1_path, file2_path, result_file_path):
    df = pd.read_excel(file1_path)
    df_HJ = pd.read_excel(file2_path) 
    df_HJ.columns
    df= df.drop(index=df.index[:2])
    df.columns=df.iloc[0]
    df=df.drop(2)
    df=df[['PI / CC','Cost Code Description','Qty Per', 'Labor','Equip', 'Other','$ Period','Period Var']]
    cost_code_col = 'PI / CC'
    complete_cost_code_col = 'CompleteCostCode'
    main_cost_code = ""
    complete_cost_codes = []
    for index, row in df.iterrows():
       cost_code = str(row[cost_code_col])
       if len(cost_code) == 9:  # Check if it's a main cost code
             main_cost_code = cost_code
             complete_cost_codes.append(main_cost_code)  # Main cost code itself
       else:
             complete_cost_code = main_cost_code + cost_code
             complete_cost_codes.append(complete_cost_code)


    df[complete_cost_code_col] = complete_cost_codes
    df['CompleteCostCode']=df['CompleteCostCode'].str.replace(' ','',regex=False)
    df_HJ=df_HJ[['Cost Code','Description','Actual Quantity','Actual Labor Cost','Actual Equipment Cost','Actual MSE Cost','Actual All Cost','All Cost Variance']]
    df=df.rename(columns={'Qty Per':'JDE_QTY', 'Labor': 'JDE_Labor', 'Equip':'JDE_Equipment','Other':'JDE_Mat_Sub', '$ Period':'JDE_Cost'})
    df.columns = df.columns.str.lower().str.replace(' ', '_')
    df_HJ.columns = df_HJ.columns.str.lower().str.replace(' ', '_')
    df.columns = df.columns.str.lower().str.replace(' ', '_')
    df_HJ.columns = df_HJ.columns.str.lower().str.replace(' ', '_')
    df_HJ=df_HJ.rename(columns={'actual_quantity':'HJ_Qty', 'actual_labor_cost' :'HJ_Labor', 'actual_equipment_cost': 'HJ_Equipment',
                        'actual_mse_cost':'HJ_Mat_Sub', 'actual_all_cost': 'HJ_Cost' })
    df.columns = df.columns.str.lower().str.replace(' ', '_')
    df_HJ.columns = df_HJ.columns.str.lower().str.replace(' ', '_')
    df=df.rename(columns={'completecostcode':'cost_code'})
    df=df[['cost_code_description', 'jde_qty',  'jde_labor', 'jde_equipment',
            'jde_mat_sub', 'jde_cost', 'period_var', 'cost_code']]
    df['cost_code'] = df['cost_code'].astype(str)
    df_HJ['cost_code'] = df_HJ['cost_code'].astype(str)
    cost_code_set=set(df_HJ['cost_code']) 
    cost_codes_with_high_difference=[]
    cost_codes_not_in_heavy_job=[] 
    df['jde_cost']= pd.to_numeric(df['jde_cost'],errors = 'coerce')  
    df_HJ['hj_cost']= pd.to_numeric(df_HJ['hj_cost'],errors='coerce')
    
    
    
    cost_codes_with_high_difference = []
    for index, row in df.iterrows():
       cost_code = row['cost_code']
       if cost_code in cost_code_set:  # Assuming cost_code_set is defined elsewhere as the set of all relevant cost codes
          df_HJ_row = df_HJ[df_HJ['cost_code'] == cost_code]
          if not df_HJ_row.empty:
             df_HJ_row = df_HJ_row.iloc[0]
             hj_cost = df_HJ_row['hj_cost']
             jde_cost = row['jde_cost']
             if hj_cost != 0:  # Ensure there is a non-zero cost to avoid division by zero
                Cost_Variance = jde_cost - hj_cost
                df_HJ.loc[df_HJ['cost_code']== cost_code,'Cost_Variance'] = Cost_Variance
                percentage_difference = abs((jde_cost - hj_cost) / hj_cost) * 100
                if percentage_difference > 10:
                   cost_codes_with_high_difference.append(cost_code)
                   
                
                   
    df_HJ['hj_labor'] = pd.to_numeric(df_HJ['hj_labor'],errors = 'coerce')
    df_HJ['hj_equipment'] = pd.to_numeric(df_HJ['hj_equipment'],errors = 'coerce')
    df_HJ[ 'hj_mat_sub'] = pd.to_numeric(df_HJ[ 'hj_mat_sub'],errors = 'coerce') 
    df['jde_labor'] = pd.to_numeric(df['jde_labor'],errors = 'coerce')
    df['jde_equipment'] = pd.to_numeric(df['jde_equipment'],errors = 'coerce')
    df['jde_mat_sub'] = pd.to_numeric(df['jde_mat_sub'],errors = 'coerce')
    
      
    high_difference_columns={}
    for cost_code in cost_codes_with_high_difference:
         df_row=df.loc[df['cost_code']==cost_code]
         df_HJ_row=df_HJ.loc[df_HJ['cost_code']==cost_code]
         if not df_row.empty and not df_HJ_row.empty:
           df_row =df_row.iloc[0]
           df_HJ_row=df_HJ_row.iloc[0]
           differences= {
               'labor':abs((df_row['jde_labor']-df_HJ_row['hj_labor'])/df_HJ_row['hj_labor'])*100 if pd.notna(df_row['jde_labor']) and pd.notna(df_HJ_row['hj_labor']) else 0,
               'equip':abs((df_row['jde_equipment']-df_HJ_row['hj_equipment'])/df_HJ_row['hj_equipment'])*100 if pd.notna(df_row['jde_equipment']) and pd.notna(df_HJ_row['hj_equipment'])else 0,
               'mse':abs((df_row['jde_mat_sub']-df_HJ_row['hj_mat_sub'])/df_HJ_row['hj_mat_sub'])*100 if pd.notna(df_row['jde_mat_sub']) and pd.notna(df_HJ_row['hj_mat_sub'])else 0
            }
           significant_differences = { k: v for k , v in differences.items() if v >= 15}
           if significant_differences:
                high_difference_columns[cost_code] = significant_differences 
                
                
                
                
    jde_cost_codes= set(df['cost_code'])
    hj_cost_codes= set(df_HJ['cost_code']) 
    unique_to_jde= jde_cost_codes - hj_cost_codes
    unique_to_hj= hj_cost_codes - jde_cost_codes 
    df['cost_code'] = df['cost_code'].str.strip().str.upper()
    df_HJ['cost_code'] = df_HJ['cost_code'].str.strip().str.upper()
    merged_df = pd.merge(df_HJ,df,on='cost_code',how='left')  
    custome_order=['cost_code', 'description', 'hj_qty','jde_qty', 'hj_labor','jde_labor', 'hj_equipment', 'jde_equipment','hj_mat_sub','jde_mat_sub', 'hj_cost','jde_cost', 'Cost_Variance']
    merged_df=merged_df[custome_order]                 
    wb = highlight_and_merge_data(df, merged_df,cost_codes_with_high_difference)
    wb = highlight_significant_columns(merged_df, high_difference_columns, wb)
    output_path = result_file_path  
    wb.save(output_path)


def highlight_and_merge_data(df, merged_df,cost_codes_with_high_difference):
       yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
       blue_fill = PatternFill(start_color="00BFFF", end_color="00BFFF", fill_type="solid")
       green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
       df['cost_code'] = df['cost_code'].str.strip().str.upper()
       merged_df['cost_code'] = merged_df['cost_code'].str.strip().str.upper()
       jde_cost_codes = set(df[df['cost_code'].apply(lambda x: len(x) != 5)]['cost_code'])
       hj_cost_codes = set(merged_df['cost_code'])
       unique_to_jde = jde_cost_codes - hj_cost_codes
       unique_to_hj = hj_cost_codes - jde_cost_codes
       new_rows = []
       for index, row in df.iterrows():
           if row['cost_code'] in unique_to_jde:
               new_row = {col: row.get(col, '') for col in merged_df.columns}
               new_row['description'] = row['cost_code_description']
               new_rows.append(new_row)
       if new_rows:
           new_rows_df = pd.DataFrame(new_rows)
           merged_df = pd.concat([merged_df, new_rows_df], ignore_index=True)
       wb = Workbook()
       ws = wb.active
       for col_idx, col_name in enumerate(merged_df.columns, 1):
           cell = ws.cell(row=1, column=col_idx, value=col_name)
           cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
       for r_idx, row in merged_df.iterrows():
           for c_idx, value in enumerate(row):
               cell = ws.cell(row=r_idx + 2, column=c_idx + 1, value=value)
               cost_code = row['cost_code']
               if cost_code in cost_codes_with_high_difference and merged_df.columns[c_idx] == 'Cost_Variance':
                   cell.fill = yellow_fill
               if cost_code in unique_to_jde and merged_df.columns[c_idx] == 'cost_code':
                   cell.fill = blue_fill
               if cost_code in unique_to_hj and merged_df.columns[c_idx] == 'cost_code':
                   cell.fill = green_fill
       return wb 


def highlight_significant_columns(merged_df, high_difference_columns, wb):
    ws = wb.active
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  
    for r_idx, row in merged_df.iterrows():
        cost_code = row['cost_code']
        if cost_code in high_difference_columns:
            differences = high_difference_columns[cost_code]
            for col in row.index:
                cell = ws.cell(row=r_idx + 2, column=row.index.get_loc(col) + 1)
                if col == 'hj_labor' and 'labor' in differences:
                    cell.fill = red_fill
                elif col == 'hj_equipment' and 'equip' in differences:
                     cell.fill = red_fill
                elif col == 'hj_mat_sub' and 'mse' in differences:
                    cell.fill = red_fill
    return wb  